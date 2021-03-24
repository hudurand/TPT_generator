import pandas as pd
import numpy as np
import re
from collections import OrderedDict
from pathlib import Path
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from timeit import default_timer as timer
import logging

from .db_fetcher import TPTFetcher
from .data_bucket import DataBucket
from .constants import IN18, FIELDS

class TPTGenerator():
    """
    Factory object to generate a TPT report for a shareclass or all specified 
    shareclasses of a subfund at a given date.
   
    It relies on the Data_Bucket object to feed it the data required to fill
    the report (and perform minor operation on the data?) to fill the columns
    correctly.

    Args:
        date (pd.DateTime): reporting date
        client (str): client the report must be generated for
        shareclass_isin (str): isin code of the shareclass the report must \
            be generated for.
    """

    def __init__(self,
                 date,
                 client,
                 output_dir,
                 source_dir,
                 sym_adj=0,
                 shareclass_isin=None):
        """
        Initialises report-specific attributes helper objects.
        """
        # Initialise logger object
        self.logger = logging.getLogger(__name__)
        self.logger.info("Initialising Generator...")

        # set in/out directories (in should be removed in future)
        self.source_dir = Path(source_dir)
        self.output_dir = Path(output_dir)
        
        # set generation constants
        self.sym_adj = sym_adj
        self.fields = FIELDS # set as attribute or global constant?
        self.IN18 = IN18 # idem?

        # instanciate DataBucket depending on production mode 
        # (single or multiple shareclasses)
        if shareclass_isin:
            self.data_bucket = DataBucket(date,
                                           client,
                                           source_dir,
                                           shareclass_isin)

            self.create_empty_report()
        else:
            self.data_bucket = DataBucket(date,
                                          client,
                                          source_dir)
            self.report = None

        self.logger.info('Generator initialiased')
        self.logger.debug(self)

    def __repr__(self):
        """
        overloads __repr__ function for logging and easier debugging.
        """
        if self.report:
            n = self.report.shape[0]
        else:
            n = self.report

        return f"""
    Reporting date:          {self.data_bucket.date}
    Client:                  {self.data_bucket.client}
    Shareclass isin:         {self.data_bucket.shareclass_isin} 
    number of instruments:   {n}
        """

    def generate(self, shareclass_isin=None):
        """
        Generates the report by calling the filling methods for all columns.
        """
        
        assert (self.data_bucket.shareclass_isin
                or shareclass_isin), "no shareclass to produce specified."
        
        self.logger.info("Begin generation")

        if isinstance(shareclass_isin, list):
            self.logger.info("Generating from list of shareclass")
            self.logger.debug(shareclass_isin)

            for isin in shareclass_isin:
                #start = timer()
                # set working isin
                self.data_bucket.update(isin)
                
                # initialise empty dataframe as report placeholder
                self.create_empty_report()
                
                # trigger generation of report
                self.fill_report()

                # save generated report
                self.output_excel()
                #end = timer()
                #print(end - start)

        else:
            
            self.logger.info("generating a single shareclass")
            self.logger.debug(shareclass_isin)

            if isinstance(shareclass_isin, str):
                # set working isin if not specified at init
                self.data_bucket.update(shareclass_isin)
                self.create_empty_report()
            
            # trigger generation of report
            self.fill_report()

            # save generated report
            self.output_excel()

    def fill_report(self):
        """
        Fills-in all data into the 
        """

        self.logger.info("Filling report")

        # call filling method for each column
        for field in self.fields:
            #print(f"fill_column_{field}...")
            self.logger.debug(f"Fill_column_{field}")

            getattr(self, f"fill_column_{field}")()

    def create_empty_report(self):
        """
        Creates an empty pandas dataframe to hold the TPT report to generate.
        """
        self.report = pd.DataFrame(index=self.data_bucket.get_instruments().index, columns=self.fields.values(), dtype=object)
    
    def output_excel(self):
        """
        Saves the generated TPT report to an excel file using the AO template.
        """

        self.logger.info("writing excel")
        self.logger.debug(self)

        client = self.data_bucket.client
        isin = self.data_bucket.shareclass_isin
        date = self.data_bucket.date

        # open template excel
        template_file_name = 'AO_TPT_V5.0_Template.xlsx'
        output_file_name = f"AO_TPT_V5.0_{client}_{isin}_{date}.xlsx"
        template = openpyxl.load_workbook(self.source_dir / template_file_name)
        report = template.get_sheet_by_name('Report')
        rows = dataframe_to_rows(self.report, index=False)

        # map dataframe columns to excel columns 
        column_map = {}
        for row_idx, row in enumerate(rows):
            if row_idx == 0:
                assert report.max_column == len(row), "Number of columns in report and template are different."
                for col_idx_pd, column_name in enumerate(row):
                    for i in range(len(row)):
                        if report.cell(row=1, column=i+1).value == column_name:
                            #print(column_name, report.cell(row=1, column=i+1).value)
                            column_map[col_idx_pd] = i+1
                    assert col_idx_pd in column_map.keys(), f"Missing {column_name} in template"
                    assert report.cell(row=1, column=column_map[col_idx_pd]).value == row[col_idx_pd]
                
            else:
                for col_idx, value in enumerate(row):
                    if value == "nan":
                        report.cell(row=row_idx+1, column=column_map[col_idx], value="")
                        report.cell(row=row_idx+1, column=column_map[col_idx]).alignment = Alignment(horizontal='center')
                    else:
                        report.cell(row=row_idx+1, column=column_map[col_idx], value=value)
                        report.cell(row=row_idx+1, column=column_map[col_idx]).alignment = Alignment(horizontal='center')

        # save produced report
        template.save(self.output_dir / output_file_name)

#    def check_required(self, required_fields):
#        """
#        Check if the fields given as input are filled and fill them if necessary.
#        """
#        for field in required_fields:
#            if self.report[self.fields[field]].isnull().values.all():
#                #print(f"fill_column_{field}")
#                getattr(self, f"fill_column_{field}")()

    def fill_instrument_info(self, info):
        """
        Abstracting method to fill fixed instruments speficific informations.
        """

        self.report[info].update(self.data_bucket.get_instruments_infos(info=info))
    
    
    def fill_scr(self, submodule):
        """
        Abstracting method to fill SCR computed values.
        """

        self.report[submodule].update(self.data_bucket.get_scr_results(submodule))

    def fill_column_1(self):
        """
        Fills column "1_Portfolio identifying data".
        
        **Definition:** isin code of the shareclass  
        **methodology:** provided by config
        """
        self.report.loc[:,self.fields["1"]] = self.data_bucket.shareclass_isin
    
    def fill_column_2(self):
        """
        Fills column "2_Type of identification code for the fund share or portfolio".
        
        **Definition:** codification chosen to identify the shareclass  
        **methodology:** reported from database
        """
        self.report.loc[:,self.fields["2"]] = \
            int(self.data_bucket.get_shareclass_infos("type_tpt"))

    def fill_column_3(self):
        """
        Fills column "3_Portfolio_name".
        
        **Definition:** definition: name of the shareclass  
        **methodology:** reported from database
        """
        self.report.loc[:,self.fields["3"]] = \
            self.data_bucket.get_shareclass_infos("shareclass_name")
    
    def fill_column_4(self):
        """
        Fills column "4_Portfolio_currency_(B)".
        
        **Definition:** valuation currency of the portfolio  
        **methodology:** reported from database
        """
        self.report.loc[:,self.fields["4"]] = \
            self.data_bucket.get_shareclass_infos("shareclass_currency")

    def fill_column_5(self):
        """
        Fills column "5_Net asset valuation of the portfolio or the share class in portfolio currency".
        
        **Definition:** NAV of the shareclass  
        **methodology:** reported from database
        """
        self.report.loc[:,self.fields["5"]] = \
            self.data_bucket.get_shareclass_nav(info="shareclass_total_net_asset_sc_curr")

    def fill_column_6(self):
        """
        Fills column "6_Valuation date".
        
        **Definition:** date of valuation  
        **methodology:** reported from database
        """
        self.report.loc[:,self.fields["6"]] = \
            self.data_bucket.get_shareclass_nav("nav_date")

    def fill_column_7(self):
        """
        Fills column "7_Reporting date".
        
        **Definition:** date of reporting  
        **methodology:** provided by config
        """
        self.report.loc[:,self.fields["7"]] = \
            self.data_bucket.date.strftime('%Y-%m-%d')

    def fill_column_8(self):
        """
        Fills column "8_Share price".
        
        **Definition:** price of one share of the shareclass  
        **methodology:** reported from database
        """
        self.report.loc[:,self.fields["8"]] = \
            self.data_bucket.get_shareclass_nav("share_price")

    def fill_column_8b(self):
        """
        Fills column "8b_Total number of shares".
        
        **Definition:** number of share emmited for the shareclass  
        **methodology:** reported from database
        """
        self.report.loc[:,self.fields["8b"]] = \
            self.data_bucket.get_shareclass_nav("outstanding_shares")

    def fill_column_9(self):
        """
        Fills column "9_% cash".
        
        **Definition:** percentage of cash in the shareclass TNA  
        **methodology:** computed from portfolio
        """
        # sum of "XT72" CIC code of the instrument divided by shareclass MV
        # TODO: replace by running bucket at init
        _ = self.data_bucket.get_processing_data("distribution")
        self.report[self.fields["9"]] = self.data_bucket.get_shareclass_infos(info="cash") \
            / self.data_bucket.get_shareclass_nav(info="shareclass_total_net_asset_sc_curr")
    
    def fill_column_10(self):
        """
        Fills column "10_Portfolio Modified Duration"
        
        **Definition:** weighted average modified duration of portfolio positions  
        **methodology:** computed from processed data  
        **require:**  
            [market exposure](processor#compute_ME)  
            [shareclass TNA](data_bucket#getshareclassnav)  
            [modified duration to next option exercise date]() 
        """
        product = self.data_bucket.get_processing_data("ME") \
                  / self.data_bucket.get_shareclass_nav(info="shareclass_total_net_asset_sf_curr") \
                  * self.data_bucket.get_instruments_infos(info=self.fields["91"])
        
        if product.isnull().values.all():
            val = np.nan
        else:
            val = product.sum()

        self.report[self.fields["10"]] = val
        
    def fill_column_11(self):
        """
        Fills column "11_Complete_SCR_delivery"
        
        **Definition:** Y/N, Y: SCR has been computed (col. 97-105)  
        **methodology:** provided by config
        """
        self.report[self.fields["11"]] = "Y"

    def fill_column_12(self):
        """
        Fills column "12_CIC_code_of_the_instrument"
        
        **Definition:** CIC Code (Complementary Identification Code)
        **methodology:** reported from database
        """
        self.fill_instrument_info(self.fields["12"])

    def fill_column_13(self):
        self.fill_instrument_info(self.fields["13"])

    def fill_column_14(self):
        self.report[self.fields["14"]] = self.report.index

    def fill_column_15(self):
        self.fill_instrument_info(self.fields["15"])

    def fill_column_16(self):
        self.fill_instrument_info(self.fields["16"])

    def fill_column_17(self):
        self.fill_instrument_info(self.fields["17"])
    
    def fill_column_17b(self):
        # Column not used in TPT
        pass

    def fill_column_18(self):
        column_18 = self.data_bucket.get_instruments("quantity_nominal") \
                    * self.data_bucket.get_distribution_weight()

        pattern = '|'.join([f"..{code}" for code in self.IN18])

        condition = (self.data_bucket.get_processing_data(self.fields["12"]).str.match(pattern) \
                    | ((self.data_bucket.get_processing_data(self.fields["12"]).str[2:] == "22") \
                    & (self.data_bucket.get_processing_data("QN").round(0) == 1.0)))

        column_18.where(condition,
                        np.nan,
                        inplace=True)

        self.report[self.fields["18"]].update(column_18)

    def fill_column_19(self):
        column_19 = self.data_bucket.get_instruments("quantity_nominal") \
                      * self.data_bucket.get_distribution_weight()
        
        pattern = '|'.join([f"..{code}" for code in self.IN18])

        condition = ~((self.data_bucket.get_processing_data(self.fields["12"]).str.match(pattern)) \
                    | ((self.data_bucket.get_processing_data(self.fields["12"]).str[2:] == "22") \
                    & ~(self.data_bucket.get_processing_data("QN").round(0) == 100.0)))

        column_19.where(condition,
                       np.nan,
                       inplace=True)

        self.report[self.fields["19"]].update(column_19)

    def fill_column_20(self):
        self.fill_instrument_info(self.fields["20"])

    def fill_column_21(self):
        self.fill_instrument_info(self.fields["21"])

    def fill_column_22(self):

        column_22 = self.data_bucket.get_instruments("market_value_asset") \
                    * self.data_bucket.get_distribution_weight()

        self.report[self.fields["22"]].update(column_22)
        
    def fill_column_23(self):
        column_23 = self.data_bucket.get_instruments("clear_value_asset") \
                    * self.data_bucket.get_distribution_weight()

        self.report[self.fields["23"]].update(column_23)

    def fill_column_24(self):
        column_24 = self.data_bucket.get_processing_data("valuation weight") \
                    * self.data_bucket.get_shareclass_nav("shareclass_total_net_asset_sc_curr")

        self.report[self.fields["24"]].update(column_24)

    def fill_column_25(self):
        
        column_25 = self.data_bucket.get_instruments("clear_value_fund") \
                    * self.data_bucket.get_processing_data("distribution weight") \
                    * self.data_bucket.get_shareclass_nav("rate")

        self.report[self.fields["25"]].update(column_25.fillna(0))

    def fill_column_26(self):
        column_26 = self.data_bucket.get_processing_data("valuation weight")

        self.report[self.fields["26"]].update(column_26)

    def fill_column_27(self):
        column_27 = self.data_bucket.get_processing_data("ME") \
                    * self.data_bucket.get_instruments("clear_value_asset") \
                    / self.data_bucket.get_instruments("clear_value_fund")

        self.report[self.fields["27"]].update(column_27)

    def fill_column_28(self):
        self.report[self.fields["28"]] = self.data_bucket.get_processing_data("ME") \
                                             * self.data_bucket.get_shareclass_nav(info="shareclass_total_net_asset_sc_curr") \
                                             / self.data_bucket.get_shareclass_nav(info="shareclass_total_net_asset_sf_curr")


    def fill_column_29(self):
        pass

    def fill_column_30(self):
        self.report[self.fields["30"]] = self.data_bucket.get_processing_data("ME") \
                                             / self.data_bucket.get_shareclass_nav(info="shareclass_total_net_asset_sf_curr")

    def fill_column_31(self):
        pass

    def fill_column_32(self):
        self.fill_instrument_info(self.fields["32"])

    def fill_column_33(self):
        self.fill_instrument_info(self.fields["33"])

    def fill_column_34(self):
        self.fill_instrument_info(self.fields["34"])

    def fill_column_35(self):
        self.fill_instrument_info(self.fields["35"])

    def fill_column_36(self):
        self.fill_instrument_info(self.fields["36"])

    def fill_column_37(self):
        self.fill_instrument_info(self.fields["37"])

    def fill_column_38(self):
        self.fill_instrument_info(self.fields["38"])

    def fill_column_39(self):
        self.fill_instrument_info(self.fields["39"])
        self.report[self.fields["39"]] = pd.to_datetime(self.report[self.fields["39"]]).dt.date
        self.report[self.fields["39"]] = self.report[self.fields["39"]].astype('str')
        self.report[self.fields["39"]].replace({"2099-12-31" : "9999-12-31"}, inplace=True)
        self.report[self.fields["39"]].replace("NaT", np.nan, inplace=True)

    def fill_column_40(self):
        self.fill_instrument_info(self.fields["40"])

    def fill_column_41(self):
        self.fill_instrument_info(self.fields["41"])

    def fill_column_42(self):
        self.fill_instrument_info(self.fields["42"])

    def fill_column_43(self):
        self.fill_instrument_info(self.fields["43"])
        self.report[self.fields["43"]] = pd.to_datetime(self.report[self.fields["43"]]).astype('str')
        self.report[self.fields["43"]].replace("NaT", np.nan, inplace=True)

    def fill_column_44(self):
        self.fill_instrument_info(self.fields["44"])

    def fill_column_45(self):
        self.fill_instrument_info(self.fields["45"])

    def fill_column_46(self):
        self.fill_instrument_info(self.fields["46"])

    def fill_column_47(self):
        self.fill_instrument_info(self.fields["47"])

    def fill_column_48(self):
        self.report[self.fields["48"]] = 9 
        self.report[self.fields["48"]].where(
            self.data_bucket.get_instruments_infos(info=self.fields["47"]).isnull(),
            1,
            inplace=True)

    def fill_column_49(self):
        self.fill_instrument_info(self.fields["49"])

    def fill_column_50(self):
        self.fill_instrument_info(self.fields["50"])

    def fill_column_51(self):
        self.report[self.fields["51"]] = 9
        self.report[self.fields["51"]].where(
            self.data_bucket.get_instruments_infos(info=self.fields["50"]).isnull(),
            1,
            inplace=True)

    def fill_column_52(self):
        self.fill_instrument_info(self.fields["52"])

    def fill_column_53(self):
        self.fill_instrument_info(self.fields["53"])

    def fill_column_54(self):
        self.fill_instrument_info(self.fields["54"])
        self.report[self.fields["54"]].where(
            self.report[self.fields["54"]].str.match("K..."),
            self.report[self.fields["54"]].str.get(0),
            inplace=True)

    def fill_column_55(self):
        self.fill_instrument_info(self.fields["55"])

    def fill_column_56(self):
        self.fill_instrument_info(self.fields["56"])

    def fill_column_57(self):
        self.fill_instrument_info(self.fields["57"])

    def fill_column_58(self):
        self.fill_instrument_info(self.fields["58"])

    def fill_column_58b(self):
        self.fill_instrument_info(self.fields["58b"])

    def fill_column_59(self):
        self.fill_instrument_info(self.fields["59"])

    def fill_column_60(self):
        self.fill_instrument_info(self.fields["60"])

    def fill_column_61(self):
        self.fill_instrument_info(self.fields["61"])
        
    def fill_column_62(self):
        self.fill_instrument_info(self.fields["62"])

    def fill_column_63(self):
        self.fill_instrument_info(self.fields["63"])

    def fill_column_64(self):
        self.fill_instrument_info(self.fields["64"])

    def fill_column_65(self):
        self.fill_instrument_info(self.fields["65"])
    
    def fill_column_67(self):
        self.fill_instrument_info(self.fields["67"])

    def fill_column_68(self):
        self.fill_instrument_info(self.fields["68"])

    def fill_column_69(self):
        self.fill_instrument_info(self.fields["69"])

    def fill_column_70(self):
        self.fill_instrument_info(self.fields["70"])

    def fill_column_71(self):
        self.fill_instrument_info(self.fields["71"])

    def fill_column_72(self):
        self.fill_instrument_info(self.fields["72"])

    def fill_column_73(self):
        self.fill_instrument_info(self.fields["73"])

    def fill_column_74(self):
        self.fill_instrument_info(self.fields["74"])

    def fill_column_75(self):
        self.fill_instrument_info(self.fields["75"])

    def fill_column_76(self):
        self.fill_instrument_info(self.fields["76"])

    def fill_column_77(self):
        self.fill_instrument_info(self.fields["77"])

    def fill_column_78(self):
        self.fill_instrument_info(self.fields["78"])

    def fill_column_79(self):
        self.fill_instrument_info(self.fields["79"])

    def fill_column_80(self):
        self.fill_instrument_info(self.fields["80"])

    def fill_column_81(self):
        self.fill_instrument_info(self.fields["81"])

    def fill_column_82(self):
        self.fill_instrument_info(self.fields["82"])

    def fill_column_83(self):
        self.fill_instrument_info(self.fields["83"])

    def fill_column_84(self):
        self.fill_instrument_info(self.fields["84"])

    def fill_column_85(self):
        self.fill_instrument_info(self.fields["85"])

    def fill_column_86(self):
        self.fill_instrument_info(self.fields["86"])

    def fill_column_87(self):
        self.fill_instrument_info(self.fields["87"])

    def fill_column_88(self):
        self.fill_instrument_info(self.fields["88"])

    def fill_column_89(self):
        self.fill_instrument_info(self.fields["89"])

    def fill_column_90(self):
        self.fill_instrument_info(self.fields["90"])

    def fill_column_91(self):
        self.fill_instrument_info(self.fields["91"])

    def fill_column_92(self):
        self.fill_instrument_info(self.fields["92"])

    def fill_column_93(self):
        self.fill_instrument_info(self.fields["93"])

    def fill_column_94(self):
        self.fill_instrument_info(self.fields["94"])

    def fill_column_94b(self):
        self.fill_instrument_info(self.fields["94b"])

    def fill_column_95(self):
        pass
    
    def fill_column_97(self):
        self.fill_scr(self.fields["97"])

    def fill_column_98(self):
        self.fill_scr(self.fields["98"])

    def fill_column_99(self):
        self.fill_scr(self.fields["99"])

    def fill_column_100(self):
        self.fill_scr(self.fields["100"])
    
    def fill_column_101(self):
        pass

    def fill_column_102(self):
        self.fill_scr(self.fields["102"])
    
    def fill_column_103(self):
        pass
    
    def fill_column_104(self):
        pass

    def fill_column_105(self):
        pass

    def fill_column_105a(self):
        self.fill_scr(self.fields["105a"])

    def fill_column_105b(self):
        self.fill_scr(self.fields["105b"])

    def fill_column_106(self):
        pass

    def fill_column_107(self):
        pass

    def fill_column_108(self):
        pass

    def fill_column_109(self):
        pass

    def fill_column_110(self):
        pass

    def fill_column_111(self):
        pass
    
    def fill_column_112(self):
        pass

    def fill_column_113(self):
        pass

    def fill_column_114(self):
        self.report[self.fields["114"]] = \
            self.data_bucket.get_instruments_infos(info=self.fields["13"]).where(
                self.data_bucket.get_instruments_infos(info=self.fields["13"]) != 0)

    def fill_column_115(self):
        self.report[self.fields["115"]] = \
            self.data_bucket.get_subfund_infos("subfund_lei")

    def fill_column_116(self):
        self.report[self.fields["116"]] = \
            9 if not self.data_bucket.get_subfund_infos("subfund_lei") else 1

    def fill_column_117(self):
        self.report.loc[:,self.fields["117"]] = \
            self.data_bucket.get_subfund_infos("subfund_name")

    def fill_column_118(self):
        self.report.loc[:,self.fields["118"]] = \
            self.data_bucket.get_subfund_infos("subfund_nace")

    def fill_column_119(self):
        self.report.loc[:,self.fields["119"]] = \
            self.data_bucket.get_fund_infos("fund_issuer_group_code")

    def fill_column_120(self):
        self.report[self.fields["120"]] = \
            9 if not self.data_bucket.get_fund_infos("fund_issuer_group_code") else 1

    def fill_column_121(self):
        self.report.loc[:,self.fields["121"]] = \
            self.data_bucket.get_fund_infos("fund_name")

    def fill_column_122(self):
        self.report.loc[:,self.fields["122"]] = \
            self.data_bucket.get_fund_infos("fund_country")

    def fill_column_123(self):
        self.report.loc[:,self.fields["123"]] = \
            self.data_bucket.get_subfund_infos("subfund_cic")

    def fill_column_123a(self):
        pass
    #    self.report.loc[:,self.fields["123a"]] = \
    #        self.data_bucket.get_fund_infos("depositary_country")

    def fill_column_124(self):
        self.fill_column_10()

        self.report[self.fields["124"]] = self.report[self.fields["10"]]

    def fill_column_125(self):
        column_125 = self.data_bucket.get_instruments("accrued_asset") \
                     * self.data_bucket.get_instruments("market_value_asset") \
                     * self.data_bucket.get_distribution_weight() \
                     / self.data_bucket.get_instruments("market_value_asset")

        self.report[self.fields["125"]].update(column_125)
        self.report[self.fields["125"]].fillna(0, inplace=True)

    def fill_column_126(self):
        column_126 = self.data_bucket.get_instruments("accrued_fund") \
                     * self.data_bucket.get_processing_data("valuation weight") \
                    * self.data_bucket.get_shareclass_nav("shareclass_total_net_asset_sc_curr") \
                     / self.data_bucket.get_instruments("market_value_fund") 

        self.report[self.fields["126"]].update(column_126)
        self.report[self.fields["126"]].fillna(0, inplace=True)

    def fill_column_127(self):
        self.fill_instrument_info(self.fields["127"])
        self.report[self.fields["127"]].replace("-", np.nan, inplace=True)

    def fill_column_128(self):
        self.fill_instrument_info(self.fields["128"])
        self.report[self.fields["128"]].replace("-", np.nan, inplace=True)
    
    def fill_column_129(self):
        pass

    def fill_column_130(self):
        pass

    def fill_column_131(self):
        column_131 = self.data_bucket.get_processing_data(self.fields["131"])

        self.report[self.fields["131"]].update(column_131)

    def fill_column_132(self):
        pass

    def fill_column_133(self):
        pass
        #self.report.loc[:,self.fields["133"]] = \
        #    self.data_bucket.get_fund_infos("depositary_name")

    def fill_column_134(self):
        pass

    def fill_column_135(self):
        pass

    def fill_column_136(self):
        pass

    def fill_column_137(self):
        self.fill_instrument_info(self.fields["137"])

    def fill_column_1000(self):
        self.report[self.fields["1000"]] = "V5.0"